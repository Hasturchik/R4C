import datetime
from openpyxl import Workbook
from django.db.models import Count
from .models import Robot
def create_excel_report():

    end_date = datetime.datetime.now()
    start_date = end_date - datetime.timedelta(days=7)

    robots = Robot.objects.filter(created__range=[start_date, end_date]) \
        .values('model', 'version') \
        .annotate(count=Count('id')) \
        .order_by('model', 'version')

    if not robots:
        return None
    
    wb = Workbook()

    models = set(robot['model'] for robot in robots)
    for model in models:
        ws = wb.create_sheet(title=model)
        ws.append(['Модель', 'Версия', 'Количество за неделю'])

        for robot in robots:
            if robot['model'] == model:
                ws.append([robot['model'], robot['version'], robot['count']])

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    return wb
